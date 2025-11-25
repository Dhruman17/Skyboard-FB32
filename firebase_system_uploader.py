#!/usr/bin/env python3
"""
Firebase System Uploader - Creates systems in Firestore from Excel file
Uses Application Default Credentials (no service account key required)
"""

import sys
import os
import argparse
from pathlib import Path
from datetime import datetime
import firebase_admin
from firebase_admin import credentials, firestore

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)


def read_excel_file(excel_path):
    """Read systems from Excel file"""
    systems = []
    print(f"📖 Reading systems from: {excel_path}")

    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        # Get header row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(cell.value.strip())

        print(f"📋 Found columns: {', '.join(headers)}")

        # Validate required columns
        required_columns = ['serialNumber', 'systemName', 'systemLocation', 'type', 'userId', 'LightOn', 'LightOff']
        missing_columns = [col for col in required_columns if col not in headers]

        if missing_columns:
            print(f"❌ Missing required columns: {', '.join(missing_columns)}")
            return None

        # Read data rows (skip header)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue

            # Create dictionary from row
            system_data = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    system_data[header] = row[idx]

            # Validate required fields
            serial_number = system_data.get('serialNumber')
            system_name = system_data.get('systemName')
            system_location = system_data.get('systemLocation')
            system_type = system_data.get('type')
            user_id = system_data.get('userId')
            light_on = system_data.get('LightOn')
            light_off = system_data.get('LightOff')

            if not all([serial_number, system_name, system_location, system_type, user_id, light_on, light_off]):
                print(f"⚠️  Row {row_num}: Missing required fields - skipping")
                continue

            # Convert serial number to string and strip whitespace
            serial_number = str(serial_number).strip()

            # Convert time values to strings
            if isinstance(light_on, datetime):
                light_on = light_on.strftime("%H:%M")
            else:
                light_on = str(light_on).strip()

            if isinstance(light_off, datetime):
                light_off = light_off.strftime("%H:%M")
            else:
                light_off = str(light_off).strip()

            system_data['serialNumber'] = serial_number
            system_data['systemName'] = str(system_name).strip()
            system_data['systemLocation'] = str(system_location).strip()
            system_data['type'] = str(system_type).strip()
            system_data['userId'] = str(user_id).strip()
            system_data['LightOn'] = light_on
            system_data['LightOff'] = light_off

            systems.append(system_data)
            print(f"✅ Row {row_num}: {serial_number} - {system_name}")

        print(f"📊 Total systems to upload: {len(systems)}")
        return systems

    except FileNotFoundError:
        print(f"❌ Excel file not found: {excel_path}")
        return None
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        import traceback
        traceback.print_exc()
        return None


def check_existing_system(db, serial_number):
    """Check if system already exists in Firestore"""
    try:
        doc_ref = db.collection('Systems').document(serial_number)
        doc = doc_ref.get()
        return doc.exists
    except Exception as e:
        print(f"⚠️  Error checking system {serial_number}: {e}")
        return False


def parse_time_to_timestamp(time_string):
    """Convert time string (HH:MM) to today's datetime timestamp"""
    from datetime import datetime, time

    try:
        # Parse the time string
        if isinstance(time_string, str):
            hour, minute = map(int, time_string.split(':'))
        else:
            return None

        # Create a datetime for today with the specified time
        today = datetime.now().date()
        dt = datetime.combine(today, time(hour, minute))

        return dt
    except Exception as e:
        print(f"⚠️  Error parsing time '{time_string}': {e}")
        return None


def create_system_in_firestore(db, system_data, skip_existing=False):
    """Create or update system in Firestore"""
    serial_number = system_data['serialNumber']

    try:
        doc_ref = db.collection('Systems').document(serial_number)

        # Check if system exists
        if skip_existing and check_existing_system(db, serial_number):
            print(f"⏭️  System {serial_number} already exists - skipping")
            return 'skipped'

        # Parse light on/off times to timestamps
        light_on_timestamp = parse_time_to_timestamp(system_data['LightOn'])
        light_off_timestamp = parse_time_to_timestamp(system_data['LightOff'])

        # Prepare Firestore document matching exact field structure
        firestore_data = {
            # Basic system info
            'serialNumber': system_data['serialNumber'],
            'systemName': system_data['systemName'],
            'systemLocation': system_data['systemLocation'],
            'type': system_data['type'],
            'systemId': system_data['userId'],  # Maps userId to systemId

            # Light control fields
            'Light_Interval_On_Time': light_on_timestamp,
            'Light_Interval_Off_Time': light_off_timestamp,
            'Light_Master_Switch': False,  # Default to off
            'Light_Time_Cycle_Switch': True,  # Default to on (time-based control)
            'time_cycle_switch_flag': True,  # Default to on

            # Environmental sensors (default values)
            'co2': 0,
            'humidity': 0.0,
            'temperature': 0.0,

            # System status
            'systemStatus': False,  # Default to offline until first connection
            'lastSeen': None,  # Will be updated when device connects
            'environmental_updated': None,  # Will be updated when device sends data

            # Firmware fields
            'version': 0.0,  # Initial version
            'firmware_url': '',  # Empty initially, set during OTA

            # Configuration
            'useCapacitive': True,  # Default to using capacitive sensors
        }

        # Create or update the document
        doc_ref.set(firestore_data, merge=True)

        exists = check_existing_system(db, serial_number)
        action = "Updated" if exists else "Created"
        print(f"✅ {action} system: {serial_number} - {system_data['systemName']}")
        return 'success'

    except Exception as e:
        print(f"❌ Failed to create system {serial_number}: {e}")
        import traceback
        traceback.print_exc()
        return 'failed'


def export_failed_systems(failed_systems, output_path):
    """Export failed systems to a new Excel file"""
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Failed Systems"

        if not failed_systems:
            return

        # Write headers
        headers = list(failed_systems[0].keys())
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)

        # Write data
        for row_num, system in enumerate(failed_systems, start=2):
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=row_num, column=col_num, value=system.get(header))

        workbook.save(output_path)
        print(f"💾 Failed systems exported to: {output_path}")

    except Exception as e:
        print(f"⚠️  Could not export failed systems: {e}")


def main():
    parser = argparse.ArgumentParser(
        description='Firebase System Uploader (No service account key required)',
        epilog='Note: Authenticate first with: gcloud auth application-default login'
    )
    parser.add_argument('--excel', default='System_Uploads.xlsx', help='Path to Excel file (default: System_Uploads.xlsx)')
    parser.add_argument('--project-id', default='skyacres-marketplace', help='Firebase project ID')
    parser.add_argument('--skip-existing', action='store_true', help='Skip systems that already exist in Firestore')
    parser.add_argument('--dry-run', action='store_true', help='Show what would be uploaded without uploading')
    parser.add_argument('--yes', action='store_true', help='Skip confirmation prompt')
    parser.add_argument('--export-failed', action='store_true', help='Export failed systems to Excel file')

    args = parser.parse_args()

    # Validate Excel file exists
    if not os.path.exists(args.excel):
        print(f"❌ Excel file not found: {args.excel}")
        sys.exit(1)

    print("=" * 60)
    print("Firebase System Uploader")
    print("Using Application Default Credentials")
    print("=" * 60)

    # Read systems from Excel
    systems = read_excel_file(args.excel)

    if not systems:
        print("❌ No systems to upload")
        sys.exit(1)

    # Dry run mode
    if args.dry_run:
        print("\n🧪 DRY RUN MODE - No actual upload")
        print(f"Would upload {len(systems)} systems:")
        for system in systems:
            print(f"  - {system['serialNumber']}: {system['systemName']} ({system['type']}) - User: {system['userId']}")
        sys.exit(0)

    # Initialize Firebase with Application Default Credentials
    try:
        cred = credentials.ApplicationDefault()
        firebase_admin.initialize_app(cred, {
            'projectId': args.project_id
        })
        db = firestore.client()
        print("✅ Firebase initialized successfully\n")
    except Exception as e:
        print(f"❌ Failed to initialize Firebase: {e}")
        print("\n💡 Please authenticate first with:")
        print("   gcloud auth application-default login")
        sys.exit(1)

    # Confirm upload
    if not args.yes:
        print(f"\n⚠️  About to upload {len(systems)} systems to Firestore.")
        if args.skip_existing:
            print("   (Existing systems will be skipped)")
        else:
            print("   (Existing systems will be updated)")
        confirm = input("Continue? (y/N): ").strip().lower()
        if confirm != 'y':
            print("❌ Upload cancelled")
            sys.exit(0)

    # Upload all systems
    print(f"\n🎯 Starting upload of {len(systems)} systems...")
    start_time = datetime.now()

    successful = 0
    failed = 0
    skipped = 0
    failed_systems = []

    for i, system in enumerate(systems, 1):
        print(f"\n--- System {i}/{len(systems)} ---")
        result = create_system_in_firestore(db, system, args.skip_existing)

        if result == 'success':
            successful += 1
        elif result == 'skipped':
            skipped += 1
        else:
            failed += 1
            failed_systems.append(system)

    end_time = datetime.now()
    duration = end_time - start_time

    # Print summary
    print(f"\n{'=' * 60}")
    print(f"📊 UPLOAD SUMMARY")
    print(f"{'=' * 60}")
    print(f"Total systems: {len(systems)}")
    print(f"✅ Successful: {successful}")
    if skipped > 0:
        print(f"⏭️  Skipped: {skipped}")
    print(f"❌ Failed: {failed}")
    print(f"⏱️  Duration: {duration}")
    print(f"📅 Completed: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")

    # Export failed systems if requested
    if failed > 0:
        print(f"\n⚠️  {failed} systems failed to upload.")
        if args.export_failed:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            failed_file = f"Failed_Systems_{timestamp}.xlsx"
            export_failed_systems(failed_systems, failed_file)

    if failed > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
